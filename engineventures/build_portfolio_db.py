"""
Build a normalized SQLite portfolio database from Portfolio_Data_RAW_practice.xlsx.

Reads each company tab by header name (never by column position -- handles
Fathom's reversed Pre-Money/Post-Money columns), filters out footnote/text
rows embedded in the data ranges, applies the unit/sign/scenario fixes
documented in data_quality_log, and recomputes ownership_pct_new_investor
as amount_raised_usd / post_money_usd for every row (source sheets' own
Ownership% columns are not trusted -- see data_quality_log for why).

Output: data/portfolio.db (SQLite) + data/data_quality_log.csv
"""

import csv
import os
import sqlite3
from datetime import datetime

import openpyxl

SRC = "Portfolio_Data_RAW_practice.xlsx"
os.makedirs("data", exist_ok=True)
DB_PATH = os.path.join("data", "portfolio.db")
LOG_CSV = os.path.join("data", "data_quality_log.csv")

NOW = datetime.utcnow().isoformat(timespec="seconds")

wb = openpyxl.load_workbook(SRC, data_only=True)

# ---------------------------------------------------------------------------
# Generic sheet extraction: map columns by header name, skip footnote/blank
# rows (any row lacking an Amount Raised value is not a real financing round).
# ---------------------------------------------------------------------------

def extract_rows(sheet_name, header_row, data_rows):
    ws = wb[sheet_name]
    hmap = {c.value: c.column for c in ws[header_row] if c.value is not None}
    out = []
    for r in data_rows:
        amount_col = hmap.get("Amount Raised")
        amount = ws.cell(row=r, column=amount_col).value if amount_col else None
        if amount is None:
            continue
        rec = {name: ws.cell(row=r, column=col).value for name, col in hmap.items()}
        out.append(rec)
    return out


def normalize_round_name(raw):
    raw = raw.strip()
    if raw in ("Series A-2", "Series A2"):
        return "Series A2"
    return raw


def parse_date(v):
    if isinstance(v, datetime):
        return v.date().isoformat()
    return None  # non-date text (e.g. "TBD 2027") -> not yet closed


def classify_round_type(name, pre_money):
    if pre_money is None:
        return "Convertible Note" if "bridge" in name.lower() else "SAFE"
    return "Priced Equity"


def pct(amount, post_money):
    if amount is None or post_money is None:
        return None
    return amount / post_money


def assign_round_orders(rows, key_fn):
    """Sequential order per company; rows whose key repeats (genuinely
    simultaneous/conflicting duplicates) share the same order number."""
    order = 0
    prev_key = None
    orders = []
    for rec in rows:
        key = key_fn(rec)
        if key != prev_key:
            order += 1
            prev_key = key
        orders.append(order)
    return orders


rounds = []          # final financing_rounds rows (dicts)
log = []             # final data_quality_log rows (dicts)


def add_round(company, round_name, round_order, date_closed, status, rtype,
              amount, pre, post, price, shares, own_new, own_fund,
              confidence, is_estimate, note):
    rounds.append(dict(
        company_name=company, round_name=round_name, round_order=round_order,
        date_closed=date_closed, round_status=status, round_type=rtype,
        amount_raised_usd=amount, pre_money_usd=pre, post_money_usd=post,
        price_per_share=price, shares_post_round=shares,
        ownership_pct_new_investor=own_new, ownership_pct_fund_position=own_fund,
        source_confidence=confidence, is_estimate=is_estimate, source_note=note,
    ))


def add_issue(company, round_name, issue, resolution, status):
    log.append(dict(company_name=company, round_name=round_name, issue=issue,
                     resolution=resolution, status=status, logged_at=NOW))


# ---------------------------------------------------------------------------
# Nimbus Robotics
# ---------------------------------------------------------------------------
raw = extract_rows("Nimbus Robotics", header_row=1, data_rows=range(2, 9))
orders = assign_round_orders(
    raw, lambda rec: (normalize_round_name(rec["Round"]), rec["Date"] if isinstance(rec["Date"], datetime) else None)
)
for rec, order in zip(raw, orders):
    name = normalize_round_name(rec["Round"])
    date_closed = parse_date(rec["Date"])
    status = "Closed" if date_closed else "Planned"
    amount = rec["Amount Raised"]
    pre, post = rec.get("Pre-Money"), rec.get("Post-Money")
    price, shares = rec.get("Price/Share"), rec.get("Shares Post-Round")
    rtype = classify_round_type(name, pre)

    if name == "Series B":  # unpriced IC guidance round
        add_round("Nimbus Robotics", name, order, None, "Planned", rtype,
                   amount, pre, post, None, None, None, None,
                   "needs_review", 1,
                   "IC guidance only; unpriced as of last IC update per sheet "
                   "footnote and Portfolio Notes (~$135M post, nothing signed). "
                   "price_per_share/shares_post_round/ownership left blank.")
    elif name == "Series A2":  # duplicate/conflicting 5/22/2023 rows
        alt = "4.5M raised / $55.5M post" if amount == 4_000_000 else "4.0M raised / $55.0M post"
        add_round("Nimbus Robotics", name, order, date_closed, status, rtype,
                   amount, pre, post, price, shares, pct(amount, post), None,
                   "needs_review", 1,
                   f"Duplicate/conflicting round dated 5/22/2023 (sheet had both "
                   f"'Series A-2' and 'Series A2' labels); cross-ref: other row "
                   f"shows ${alt}. Not reconciled -- see data_quality_log.")
    else:
        add_round("Nimbus Robotics", name, order, date_closed, status, rtype,
                   amount, pre, post, price, shares, pct(amount, post), None,
                   "confirmed", 0, "Tracker sheet, reconciled, no adjustments.")

add_issue("Nimbus Robotics", "Series B",
          "Series B date given as text 'TBD 2027' rather than a real date; "
          "round not yet priced per sheet footnote.",
          "Set round_status='Planned', date_closed=NULL, and left "
          "price_per_share/shares_post_round/ownership columns blank; "
          "amount/pre/post-money retained as IC guidance figures (is_estimate=1).",
          "Resolved")
add_issue("Nimbus Robotics", "Series A2",
          "Two rows both dated 5/22/2023, one labeled 'Series A-2' ($4.0M "
          "raised, $55.0M post) and one 'Series A2' ($4.5M raised, $55.5M "
          "post) -- conflicting or duplicate entries with no basis in the "
          "workbook to determine which (if either) is correct, or whether "
          "they represent sequential tranches.",
          "Not resolved -- inserted both rows sharing round_order, each "
          "flagged source_confidence='needs_review', is_estimate=1, with "
          "source_note cross-referencing the other row's figures.",
          "Open")
add_issue("Nimbus Robotics", "Bridge Note",
          "Portfolio Notes: 'nimbus robotics - bridge note closed 8/1/24, "
          "$2.0M, existing investors only'.",
          "Cross-checked against tracker (Bridge Note, 8/1/2024, $2.0M) -- "
          "matches exactly, no change needed.",
          "Resolved")
add_issue("Nimbus Robotics", "Series B",
          "Portfolio Notes: 'Nimbus Robotics Series B still not priced, IC "
          "guidance is ~$135M post but nothing signed'.",
          "Cross-checked against tracker -- confirms Series B is correctly "
          "represented as unpriced/Planned with ~$135M post-money guidance, "
          "no change needed.",
          "Resolved")

# ---------------------------------------------------------------------------
# Verdant Bio (Continue Pro-Rata Participation scenario only, rows 4-7)
# ---------------------------------------------------------------------------
raw = extract_rows("Verdant Bio", header_row=3, data_rows=range(4, 8))
orders = assign_round_orders(raw, lambda rec: (rec["Round"], rec["Date"]))
for rec, order in zip(raw, orders):
    name = rec["Round"]
    date_closed = parse_date(rec["Date"])
    amount, pre, post = rec["Amount Raised"], rec.get("Pre-Money"), rec.get("Post-Money")
    fund_pos = rec.get("Ownership %")
    rtype = classify_round_type(name, pre)

    if name == "Series C":
        add_round("Verdant Bio", name, order, None, "Planned", rtype,
                   amount, pre, post, None, None, pct(amount, post), fund_pos,
                   "needs_review", 1,
                   "Expected/target close date 2025-04-01 per Portfolio Notes "
                   "('pushed to Q2 2025' update); round not yet confirmed "
                   "closed as of last tracker update, so date_closed left "
                   "NULL and status set to Planned.")
    else:
        add_round("Verdant Bio", name, order, date_closed, "Closed", rtype,
                   amount, pre, post, None, None, pct(amount, post), fund_pos,
                   "confirmed", 0,
                   "Tracker sheet ('Continue Pro-Rata Participation' "
                   "scenario); stored Ownership% preserved as fund position "
                   "metric, see data_quality_log.")

add_issue("Verdant Bio", None,
          "Sheet stacks two scenarios in one tab: 'Continue Pro-Rata "
          "Participation' and 'Pause Investing After Series B'.",
          "Only 'Continue Pro-Rata Participation' rows inserted as "
          "realized/base-case history. 'Pause' scenario excluded entirely "
          "as a hypothetical forward-looking branch, not settled history.",
          "Resolved")
add_issue("Verdant Bio", None,
          "Stored Ownership% column (25% -> 19.2% -> 7.3% -> 5.2%, "
          "monotonically decreasing) diverges from amount_raised/post_money "
          "recompute (25% -> 23.1% -> 18.2% -> 14.9%).",
          "Recomputed ownership_pct_new_investor per the standard rule for "
          "all rows. Stored values represent Engine's own cumulative "
          "diluting fund position, not each round's new-money %, and were "
          "preserved separately in ownership_pct_fund_position.",
          "Resolved")
add_issue("Verdant Bio", "Series C",
          "Series C carries a firm date (4/1/2025) and full deal terms in "
          "the tracker, but is not confirmed closed.",
          "Set round_status='Planned', date_closed=NULL. Expected/target "
          "close date (2025-04-01, reflecting the 'pushed to Q2 2025' "
          "update in Portfolio Notes) recorded in source_note instead of "
          "date_closed.",
          "Resolved")
add_issue("Verdant Bio", "Series C",
          "Portfolio Notes: 'Verdant Bio Series C round pushed back a "
          "quarter, expected close now Q2 2025 not Q1'.",
          "Cross-checked against tracker date (4/1/2025, which is Q2) -- "
          "already reflects this update, no date change needed. (Status "
          "handled separately, see other Series C entry.)",
          "Resolved")

# ---------------------------------------------------------------------------
# Fathom Analytics
# ---------------------------------------------------------------------------
raw = extract_rows("Fathom Analytics", header_row=1, data_rows=range(2, 6))
orders = assign_round_orders(raw, lambda rec: (rec["Round"], rec["Date"]))
for rec, order in zip(raw, orders):
    name = rec["Round"]
    date_closed = parse_date(rec["Date"])
    amount, pre, post = rec["Amount Raised"], rec.get("Pre-Money"), rec.get("Post-Money")
    price = rec.get("Price/Share")
    rtype = classify_round_type(name, pre)

    if name == "Pre-Seed":
        add_round("Fathom Analytics", name, order, date_closed, "Closed", rtype,
                   amount, pre, post, price, None, pct(amount, post), None,
                   "needs_review", 0,
                   "Tracker's existing $750K figure retained; Portfolio "
                   "Notes mentions an unconfirmed 'new SAFE $1.2M dated "
                   "1/10/22' referencing the same date -- possible "
                   "duplicate/conflict, not applied. See data_quality_log.")
    else:
        add_round("Fathom Analytics", name, order, date_closed, "Closed", rtype,
                   amount, pre, post, price, None, pct(amount, post), None,
                   "confirmed", 0, "Tracker sheet, reconciled, no adjustments.")

add_issue("Fathom Analytics", "Pre-Seed",
          "Portfolio Notes mentions an unconfirmed 'new SAFE, $1.2M, dated "
          "1/10/22' -- same date as the tracker's existing Pre-Seed row, "
          "which shows $750K raised. Possible duplicate or conflicting "
          "entry for the same event.",
          "Kept the tracker's existing $750K figure (not overwritten by "
          "the unconfirmed note); flagged source_confidence='needs_review' "
          "pending deal-team confirmation.",
          "Open")
add_issue("Fathom Analytics", "Series B",
          "Portfolio Notes: 'Fathom Analytics Inc. Series B - closed June "
          "2025, $22M raised, post money ~$92M'.",
          "Cross-checked against tracker (6/18/2025, $22M raised, $92M "
          "post) -- matches, no change needed.",
          "Resolved")

# ---------------------------------------------------------------------------
# Ridgeline Materials ($000s -> multiply by 1000)
# ---------------------------------------------------------------------------
RIDGELINE_MULT = 1000
raw = extract_rows("Ridgeline Materials", header_row=1, data_rows=range(2, 7))
orders = assign_round_orders(raw, lambda rec: (rec["Round"], rec["Date"]))
for rec, order in zip(raw, orders):
    name = rec["Round"]
    date_closed = parse_date(rec["Date"])
    amount = rec["Amount Raised"]
    pre, post = rec.get("Pre-Money"), rec.get("Post-Money")

    note = "Unit conversion: source figures in $000s, multiplied by 1,000."
    if name == "Series B" and amount < 0:
        amount = abs(amount)
        note += (" Amount Raised corrected from -22,000 to +22,000 (typo "
                 "confirmed by sheet footnote and Portfolio Notes).")

    amount = amount * RIDGELINE_MULT if amount is not None else None
    pre = pre * RIDGELINE_MULT if pre is not None else None
    post = post * RIDGELINE_MULT if post is not None else None
    rtype = classify_round_type(name, pre)

    add_round("Ridgeline Materials", name, order, date_closed, "Closed", rtype,
               amount, pre, post, None, None, pct(amount, post), None,
               "confirmed", 0, note)

add_issue("Ridgeline Materials", None,
          "Sheet footnote states all dollar figures are in $000s, differing "
          "from every other sheet's raw-dollar convention.",
          "Multiplied all Ridgeline dollar fields (amount_raised_usd, "
          "pre_money_usd, post_money_usd) by 1,000 before inserting.",
          "Resolved")
add_issue("Ridgeline Materials", "Series B",
          "Amount Raised entered as -22,000 (in $000s) -- both the sheet's "
          "own footnote and Portfolio Notes confirm this is a typo.",
          "Corrected to +22,000 ($22M after unit conversion) before insert.",
          "Resolved")

# ---------------------------------------------------------------------------
# Halcyon Health
# ---------------------------------------------------------------------------
raw = extract_rows("Halcyon Health", header_row=1, data_rows=range(2, 6))
# Drop the broken $9.0M Series A row (shares=0, #DIV/0!) before assigning order,
# so it never becomes an inserted row or consumes an order slot.
raw = [r for r in raw if not (r["Round"] == "Series A" and r.get("Shares Post-Round") == 0)]
orders = assign_round_orders(raw, lambda rec: (rec["Round"], rec["Date"]))
for rec, order in zip(raw, orders):
    name = rec["Round"]
    date_closed = parse_date(rec["Date"])
    amount, pre, post = rec["Amount Raised"], rec.get("Pre-Money"), rec.get("Post-Money")
    shares = rec.get("Shares Post-Round")
    shares = None if shares in (0, None) else shares
    rtype = classify_round_type(name, pre)

    if name == "Series A":
        add_round("Halcyon Health", name, order, date_closed, "Closed", rtype,
                   amount, pre, post, None, shares, pct(amount, post), None,
                   "needs_review", 1,
                   "Two Series A rows existed in source (sheet footnote: "
                   "'not yet reconciled'); this $9.5M/15.2M-share row "
                   "retained as the only one with usable share data. "
                   "Alternate unconfirmed figure: $9.0M raised, "
                   "shares_post_round and Ownership% broken (#DIV/0!) in "
                   "source. Pending finance confirmation per Portfolio Notes.")
    elif name == "Series B":
        add_round("Halcyon Health", name, order, date_closed, "Closed", rtype,
                   amount, pre, post, None, None, pct(amount, post), None,
                   "confirmed", 0,
                   "shares_post_round unavailable in source (#DIV/0!); left "
                   "NULL rather than estimated. Amount/pre/post-money "
                   "figures reliable independent of the broken share count.")
    else:
        add_round("Halcyon Health", name, order, date_closed, "Closed", rtype,
                   amount, pre, post, None, shares, pct(amount, post), None,
                   "confirmed", 0, "Tracker sheet, reconciled, no adjustments.")

add_issue("Halcyon Health", "Series A",
          "Two conflicting Series A rows in source: $9.0M raised "
          "(shares_post_round=0, #DIV/0! ownership -- broken formula "
          "artifact) vs. $9.5M raised (complete, 15.2M shares). Sheet "
          "footnote and Portfolio Notes both flag as unreconciled, pending "
          "finance confirmation.",
          "Inserted only the $9.5M row (the only one with usable share "
          "data), flagged source_confidence='needs_review', is_estimate=1. "
          "$9.0M figure preserved here for reference, not inserted as a row.",
          "Open")
add_issue("Halcyon Health", "Series B",
          "shares_post_round was 0/#DIV/0! in source (broken formula, no "
          "real share count available).",
          "Left shares_post_round and price_per_share NULL rather than "
          "estimate a value. ownership_pct_new_investor still computed "
          "from amount_raised_usd/post_money_usd, both of which are reliable.",
          "Resolved")

# ---------------------------------------------------------------------------
# Dataset-wide issues (span multiple companies -> single consolidated row)
# ---------------------------------------------------------------------------
add_issue("All Companies", None,
          "Every sheet's stored Ownership% column uses a different, "
          "unreliable basis (e.g. Halcyon's equals prior_shares/"
          "total_shares_post_round, a cap-table stat, not deal ownership; "
          "two Halcyon rows are outright #DIV/0!).",
          "Recomputed ownership_pct_new_investor for every row as "
          "amount_raised_usd / post_money_usd; no sheet's stored "
          "Ownership% column was used directly (Verdant's stored values "
          "were preserved separately as ownership_pct_fund_position, see "
          "separate entry).",
          "Resolved")
add_issue("All Companies", None,
          "Footnote/context text rows are embedded directly inside the "
          "data ranges on several sheets (Nimbus row 11, Ridgeline row 8, "
          "Halcyon row 7, Verdant row 17), which if not filtered would "
          "become garbage financing_rounds rows with null amounts.",
          "Rows without a valid Amount Raised value were excluded from "
          "extraction; footnote text was reviewed and folded into the "
          "relevant data_quality_log entries above as context rather than "
          "inserted as rows.",
          "Resolved")

# ---------------------------------------------------------------------------
# Write to SQLite
# ---------------------------------------------------------------------------
conn = sqlite3.connect(DB_PATH)
cur = conn.cursor()
cur.executescript("""
DROP TABLE IF EXISTS data_quality_log;
DROP TABLE IF EXISTS financing_rounds;
DROP TABLE IF EXISTS companies;

CREATE TABLE companies (
    company_id INTEGER PRIMARY KEY AUTOINCREMENT,
    company_name TEXT UNIQUE NOT NULL
);

CREATE TABLE financing_rounds (
    round_id INTEGER PRIMARY KEY AUTOINCREMENT,
    company_id INTEGER NOT NULL REFERENCES companies(company_id),
    round_name TEXT NOT NULL,
    round_order INTEGER NOT NULL,
    date_closed TEXT,
    round_status TEXT NOT NULL CHECK (round_status IN ('Closed', 'Planned')),
    round_type TEXT NOT NULL CHECK (round_type IN ('SAFE', 'Convertible Note', 'Priced Equity')),
    amount_raised_usd REAL,
    pre_money_usd REAL,
    post_money_usd REAL,
    price_per_share REAL,
    shares_post_round REAL,
    ownership_pct_new_investor REAL,
    ownership_pct_fund_position REAL,
    source_confidence TEXT NOT NULL CHECK (source_confidence IN ('confirmed', 'needs_review')),
    is_estimate INTEGER NOT NULL CHECK (is_estimate IN (0, 1)),
    source_note TEXT,
    created_at TEXT NOT NULL,
    updated_at TEXT NOT NULL
);

CREATE TABLE data_quality_log (
    issue_id INTEGER PRIMARY KEY AUTOINCREMENT,
    company_name TEXT NOT NULL,
    round_name TEXT,
    issue TEXT NOT NULL,
    resolution TEXT NOT NULL,
    status TEXT NOT NULL CHECK (status IN ('Resolved', 'Open')),
    logged_at TEXT NOT NULL
);
""")

COMPANIES = ["Nimbus Robotics", "Verdant Bio", "Fathom Analytics",
             "Ridgeline Materials", "Halcyon Health"]
for name in COMPANIES:
    cur.execute("INSERT INTO companies (company_name) VALUES (?)", (name,))
company_id = {name: cid for cid, name in
              cur.execute("SELECT company_id, company_name FROM companies")}

for r in rounds:
    cur.execute("""
        INSERT INTO financing_rounds (
            company_id, round_name, round_order, date_closed, round_status,
            round_type, amount_raised_usd, pre_money_usd, post_money_usd,
            price_per_share, shares_post_round, ownership_pct_new_investor,
            ownership_pct_fund_position, source_confidence, is_estimate,
            source_note, created_at, updated_at
        ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
    """, (
        company_id[r["company_name"]], r["round_name"], r["round_order"],
        r["date_closed"], r["round_status"], r["round_type"],
        r["amount_raised_usd"], r["pre_money_usd"], r["post_money_usd"],
        r["price_per_share"], r["shares_post_round"],
        r["ownership_pct_new_investor"], r["ownership_pct_fund_position"],
        r["source_confidence"], r["is_estimate"], r["source_note"], NOW, NOW,
    ))

for i in log:
    cur.execute("""
        INSERT INTO data_quality_log (
            company_name, round_name, issue, resolution, status, logged_at
        ) VALUES (?,?,?,?,?,?)
    """, (i["company_name"], i["round_name"], i["issue"], i["resolution"],
          i["status"], i["logged_at"]))

conn.commit()

# ---------------------------------------------------------------------------
# Sanity checks
# ---------------------------------------------------------------------------
print("=" * 80)
print("SANITY CHECKS")
print("=" * 80)

missing_amount = cur.execute("""
    SELECT c.company_name, fr.round_name FROM financing_rounds fr
    JOIN companies c ON c.company_id = fr.company_id
    WHERE fr.round_status = 'Closed' AND fr.amount_raised_usd IS NULL
""").fetchall()
print(f"Closed rounds missing amount_raised_usd: {len(missing_amount)}")
for row in missing_amount:
    print("  FAIL:", row)

bad_ownership = cur.execute("""
    SELECT c.company_name, fr.round_name, fr.ownership_pct_new_investor
    FROM financing_rounds fr
    JOIN companies c ON c.company_id = fr.company_id
    WHERE fr.ownership_pct_new_investor < 0 OR fr.ownership_pct_new_investor > 1.0
""").fetchall()
print(f"ownership_pct_new_investor out of [0,1] range: {len(bad_ownership)}")
for row in bad_ownership:
    print("  FAIL:", row)

if not missing_amount and not bad_ownership:
    print("All sanity checks passed.")

# ---------------------------------------------------------------------------
# Export data_quality_log for review
# ---------------------------------------------------------------------------
rows_out = cur.execute("""
    SELECT issue_id, company_name, round_name, issue, resolution, status, logged_at
    FROM data_quality_log ORDER BY issue_id
""").fetchall()
with open(LOG_CSV, "w", newline="", encoding="utf-8") as f:
    writer = csv.writer(f)
    writer.writerow(["issue_id", "company_name", "round_name", "issue",
                      "resolution", "status", "logged_at"])
    writer.writerows(rows_out)

print("\n" + "=" * 80)
print(f"DATA QUALITY LOG ({len(rows_out)} entries) -> {LOG_CSV}")
print("=" * 80)
for row in rows_out:
    issue_id, comp, rname, issue, resolution, status, logged_at = row
    print(f"\n[{issue_id}] {comp}" + (f" / {rname}" if rname else "") + f" -- {status}")
    print(f"  ISSUE:      {issue}")
    print(f"  RESOLUTION: {resolution}")

open_count = sum(1 for r in rows_out if r[5] == "Open")
resolved_count = sum(1 for r in rows_out if r[5] == "Resolved")
print(f"\nTotal: {len(rows_out)}  |  Resolved: {resolved_count}  |  Open: {open_count}")
print(f"\nDatabase written to {DB_PATH}")

conn.close()
