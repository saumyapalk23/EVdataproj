"""
Build AND CLEAN the MySQL "portfolio" database from Portfolio_Data_RAW_practice.xlsx.

Reads each company tab by header name (never by column position -- handles
Fathom's reversed Pre-Money/Post-Money columns), filters out footnote/text
rows embedded in the data ranges, applies the unit/sign/scenario fixes
documented in data_quality_log, and recomputes ownership_pct_new_investor
as amount_raised_usd / post_money_usd for every row.

Output: data/portfolio.sql (MySQL dump, executed against MySQL by this
script). The Streamlit app (app.py) reads the resulting MySQL database
directly; nothing in this pipeline uses SQLite.
"""

import os
import tomllib
from collections import defaultdict
from datetime import datetime

import openpyxl
import pymysql
from pymysql.constants import CLIENT

# Source workbook this script reads from, and the output artifact it
# produces: a plain-text MySQL SQL dump (the version-controlled,
# human-diffable source of truth, also executed against MySQL by this
# script to actually build the database).
SRC = "Portfolio_Data_RAW_practice.xlsx"
os.makedirs("data", exist_ok=True)
SQL_DUMP_PATH = os.path.join("data", "portfolio.sql")

# Reuse the same MySQL credentials as app.py, rather than duplicating them.
with open(os.path.join(".streamlit", "secrets.toml"), "rb") as f:
    MYSQL_CREDS = tomllib.load(f)["mysql"]

# Single timestamp reused for every row inserted by this run, so a run's
# created_at/updated_at/logged_at values are all consistent with each other.
NOW = datetime.utcnow().isoformat(timespec="seconds")

# data_only=True reads cached formula *results* rather than formula text
# (several sheets have #DIV/0! ownership formulas that only make sense as values).
wb = openpyxl.load_workbook(SRC, data_only=True)

# ---------------------------------------------------------------------------
# Generic sheet extraction: map columns by header name, skip footnote/blank
# rows (any row lacking an Amount Raised value is not a real financing round).
# ---------------------------------------------------------------------------

def extract_rows(sheet_name, header_row, data_rows):
    """Read a rectangular block of a worksheet into a list of dicts, keyed by
    each column's header text (not its position) -- this is what lets the
    same code handle Fathom's Pre-Money/Post-Money columns being swapped
    relative to other sheets. Rows with no Amount Raised value are skipped
    since they're footnotes/blank rows, not real financing rounds."""
    ws = wb[sheet_name]
    # Build header-name -> column-index map from the header row.
    hmap = {c.value: c.column for c in ws[header_row] if c.value is not None}
    out = []
    for r in data_rows:
        amount_col = hmap.get("Amount Raised")
        amount = ws.cell(row=r, column=amount_col).value if amount_col else None
        if amount is None:
            continue
        # Pull every mapped column's value for this row into a dict.
        rec = {name: ws.cell(row=r, column=col).value for name, col in hmap.items()}
        out.append(rec)
    return out


def normalize_round_name(raw):
    """Collapse the two spellings of the same round ('Series A-2' and
    'Series A2') seen in Nimbus's sheet down to one canonical name."""
    raw = raw.strip()
    if raw in ("Series A-2", "Series A2"):
        return "Series A2"
    return raw


def parse_date(v):
    """Convert an Excel cell value to an ISO date string, or None if it's
    non-date placeholder text (e.g. Nimbus's Series B 'TBD 2027') -- which
    signals the round hasn't actually closed yet."""
    if isinstance(v, datetime):
        return v.date().isoformat()
    return None  # non-date text (e.g. "TBD 2027") -> not yet closed


def classify_round_type(name, pre_money):
    """Infer round_type from whether a pre-money valuation exists. No
    pre-money means the round is unpriced (SAFE, or a Convertible Note if
    the round name mentions "bridge"); a pre-money value means it's priced equity."""
    if pre_money is None:
        return "Convertible Note" if "bridge" in name.lower() else "SAFE"
    return "Priced Equity"


def pct(amount, post_money):
    """Compute ownership_pct_new_investor = amount raised / post-money.
    This is the single, consistent basis used for every row in the database,
    replacing each sheet's own (unreliable) Ownership% column."""
    if amount is None or post_money is None:
        return None
    return amount / post_money


def assign_round_orders(rows, key_fn):
    """Sequential order per company; rows whose key repeats (genuinely
    simultaneous/conflicting duplicates) share the same order number."""
    order = 0
    prev_key = None
    orders = []
    for rec in rows:
        key = key_fn(rec)
        # Only bump the order counter when the key actually changes, so
        # consecutive rows sharing a key (e.g. two conflicting entries for
        # the same round/date) are treated as one position in the sequence.
        if key != prev_key:
            order += 1
            prev_key = key
        orders.append(order)
    return orders


# The fixed list of portfolio companies, in the order they'll get
# AUTO_INCREMENT ids (1..5) when inserted into MySQL below -- this dict lets
# add_issue() (and the SQL-dump generation further down) resolve each
# company_name to its company_id without touching a database.
COMPANIES = ["Nimbus Robotics", "Verdant Bio", "Fathom Analytics",
             "Ridgeline Materials", "Halcyon Health"]
company_id = {name: idx for idx, name in enumerate(COMPANIES, start=1)}

# In-memory accumulators populated by the per-company blocks below, then
# bulk-inserted into SQLite in the "Write to SQLite" section further down.
rounds = []          # final financing_rounds rows (dicts)
log = []             # final data_quality_log rows (dicts)


def add_round(company, round_name, round_order, date_closed, status, rtype,
              amount, pre, post, price, shares, own_new, own_fund,
              confidence, is_estimate, note):
    """Append one financing_rounds row (as a dict) to the `rounds` accumulator.
    Returns the round's round_id (its 1-based position, matching the id it
    will be assigned during SQL-dump generation below) so a caller can pass
    it to add_issue() and link a log entry to this exact round."""
    rounds.append(dict(
        company_name=company, round_name=round_name, round_order=round_order,
        date_closed=date_closed, round_status=status, round_type=rtype,
        amount_raised_usd=amount, pre_money_usd=pre, post_money_usd=post,
        price_per_share=price, shares_post_round=shares,
        ownership_pct_new_investor=own_new, ownership_pct_fund_position=own_fund,
        source_confidence=confidence, is_estimate=is_estimate, source_note=note,
    ))
    return len(rounds)


def add_issue(company, round_name, issue, resolution, status, round_id=None):
    """Append one data_quality_log row (as a dict) to the `log` accumulator.
    This is the audit trail explaining every non-obvious judgment call made
    while normalizing the raw workbook data below.

    company_id is resolved from the COMPANIES lookup above (None for the
    synthetic "All Companies" dataset-wide rows, which have no single
    company to point to). round_id must be passed in explicitly -- the value
    just returned by add_round() -- since round_name alone doesn't
    disambiguate cases like Nimbus's duplicate/conflicting Series A2 rows."""
    log.append(dict(company_name=company, round_name=round_name, issue=issue,
                     resolution=resolution, status=status, logged_at=NOW,
                     company_id=company_id.get(company), round_id=round_id))


# ---------------------------------------------------------------------------
# Nimbus Robotics
# ---------------------------------------------------------------------------
# Rows 2-8 of the "Nimbus Robotics" sheet (row 1 is the header).
raw = extract_rows("Nimbus Robotics", header_row=1, data_rows=range(2, 9))
# Order rounds by (normalized name, date); two rows sharing both a name and
# date are treated as duplicate/conflicting entries for the same round slot.
orders = assign_round_orders(
    raw, lambda rec: (normalize_round_name(rec["Round"]), rec["Date"] if isinstance(rec["Date"], datetime) else None)
)
# round_ids of specific rows referenced by the add_issue() calls below.
nimbus_series_b_id = None
nimbus_series_a2_ids = []
nimbus_bridge_note_id = None
for rec, order in zip(raw, orders):
    name = normalize_round_name(rec["Round"])
    date_closed = parse_date(rec["Date"])
    status = "Closed" if date_closed else "Planned"
    amount = rec["Amount Raised"]
    pre, post = rec.get("Pre-Money"), rec.get("Post-Money")
    price, shares = rec.get("Price/Share"), rec.get("Shares Post-Round")
    rtype = classify_round_type(name, pre)

    if name == "Series B":  # unpriced IC guidance round
        # Per the sheet's own footnote and the Portfolio Notes, Series B is
        # only IC (investment committee) guidance -- nothing has actually
        # been signed, so it's inserted as Planned/unpriced with the
        # priced-round-only fields (price/shares/ownership) left blank.
        nimbus_series_b_id = add_round(
                   "Nimbus Robotics", name, order, None, "Planned", rtype,
                   amount, pre, post, None, None, None, None,
                   "needs_review", 1,
                   "IC guidance only; unpriced as of last IC update per sheet "
                   "footnote and Portfolio Notes (~$135M post, nothing signed). "
                   "price_per_share/shares_post_round/ownership_pct_new_investor "
                   "left blank at extraction, later back-computed as "
                   "projected/unconfirmed estimates -- see data_quality_log.")
    elif name == "Series A2":  # duplicate/conflicting 5/22/2023 rows
        # Two rows for the same date/round with different dollar figures --
        # both are inserted (sharing round_order) since there's no basis in
        # the source workbook to pick one as correct; each is flagged.
        alt = "4.5M raised / $55.5M post" if amount == 4_000_000 else "4.0M raised / $55.0M post"
        nimbus_series_a2_ids.append(add_round(
                   "Nimbus Robotics", name, order, date_closed, status, rtype,
                   amount, pre, post, price, shares, pct(amount, post), None,
                   "needs_review", 1,
                   f"Duplicate/conflicting round dated 5/22/2023 (sheet had both "
                   f"'Series A-2' and 'Series A2' labels); cross-ref: other row "
                   f"shows ${alt}. Not reconciled -- see data_quality_log."))
    else:
        # Ordinary, unambiguous round -- insert as-is with the standard
        # recomputed ownership_pct_new_investor.
        rid = add_round("Nimbus Robotics", name, order, date_closed, status, rtype,
                   amount, pre, post, price, shares, pct(amount, post), None,
                   "confirmed", 0, "Tracker sheet, reconciled, no adjustments.")
        if name == "Bridge Note":
            nimbus_bridge_note_id = rid

# Data-quality entries documenting the judgment calls made above, plus two
# entries that simply cross-check the sheet against separately-supplied
# Portfolio Notes text (no data changes resulted from those checks).
add_issue("Nimbus Robotics", "Series B",
          "Series B date given as text 'TBD 2027' rather than a real date; "
          "round not yet priced per sheet footnote.",
          "Set round_status='Planned', date_closed=NULL, and left "
          "price_per_share/shares_post_round/ownership columns blank; "
          "amount/pre/post-money retained as IC guidance figures (is_estimate=1).",
          "Resolved", round_id=nimbus_series_b_id)
add_issue("Nimbus Robotics", "Series A2",
          "Two rows both dated 5/22/2023, one labeled 'Series A-2' ($4.0M "
          "raised, $55.0M post) and one 'Series A2' ($4.5M raised, $55.5M "
          "post) -- conflicting or duplicate entries with no basis in the "
          "workbook to determine which (if either) is correct, or whether "
          "they represent sequential tranches.",
          "Not resolved -- inserted both rows sharing round_order, each "
          "flagged source_confidence='needs_review', is_estimate=1, with "
          "source_note cross-referencing the other row's figures.",
          "Open", round_id=nimbus_series_a2_ids[0] if nimbus_series_a2_ids else None)
add_issue("Nimbus Robotics", "Bridge Note",
          "Portfolio Notes: 'nimbus robotics - bridge note closed 8/1/24, "
          "$2.0M, existing investors only'.",
          "Cross-checked against tracker (Bridge Note, 8/1/2024, $2.0M) -- "
          "matches exactly, no change needed.",
          "Resolved", round_id=nimbus_bridge_note_id)
add_issue("Nimbus Robotics", "Series B",
          "Portfolio Notes: 'Nimbus Robotics Series B still not priced, IC "
          "guidance is ~$135M post but nothing signed'.",
          "Cross-checked against tracker -- confirms Series B is correctly "
          "represented as unpriced/Planned with ~$135M post-money guidance, "
          "no change needed.",
          "Resolved", round_id=nimbus_series_b_id)

# ---------------------------------------------------------------------------
# Verdant Bio (Continue Pro-Rata Participation scenario only, rows 4-7)
# ---------------------------------------------------------------------------
# Header is on row 3 here (not row 1); only rows 4-7 belong to the
# "Continue Pro-Rata Participation" scenario -- see the sheet-structure note below.
raw = extract_rows("Verdant Bio", header_row=3, data_rows=range(4, 8))
orders = assign_round_orders(raw, lambda rec: (rec["Round"], rec["Date"]))
verdant_series_c_id = None
for rec, order in zip(raw, orders):
    name = rec["Round"]
    date_closed = parse_date(rec["Date"])
    amount, pre, post = rec["Amount Raised"], rec.get("Pre-Money"), rec.get("Post-Money")
    fund_pos = rec.get("Ownership %")
    rtype = classify_round_type(name, pre)

    if name == "Series C":
        # Series C has full deal terms in the sheet but is not confirmed
        # closed, so it's recorded as Planned with the expected date moved
        # to source_note rather than date_closed.
        verdant_series_c_id = add_round(
                   "Verdant Bio", name, order, None, "Planned", rtype,
                   amount, pre, post, None, None, pct(amount, post), fund_pos,
                   "needs_review", 1,
                   "Expected/target close date 2025-04-01 per Portfolio Notes "
                   "('pushed to Q2 2025' update); round not yet confirmed "
                   "closed as of last tracker update, so date_closed left "
                   "NULL and status set to Planned.")
    else:
        # fund_pos (the sheet's stored Ownership% column) is preserved
        # separately as ownership_pct_fund_position -- it represents
        # Engine's cumulative diluting position, not this round's new-money %.
        add_round("Verdant Bio", name, order, date_closed, "Closed", rtype,
                   amount, pre, post, None, None, pct(amount, post), fund_pos,
                   "confirmed", 0,
                   "Tracker sheet ('Continue Pro-Rata Participation' "
                   "scenario); stored Ownership% preserved as fund position "
                   "metric, see data_quality_log.")

add_issue("Verdant Bio", None,
          "Sheet stacks two scenarios in one tab: 'Continue Pro-Rata "
          "Participation' and 'Pause Investing After Series B'.",
          "Only 'Continue Pro-Rata Participation' rows inserted as "
          "realized/base-case history. 'Pause' scenario excluded entirely "
          "as a hypothetical forward-looking branch, not settled history.",
          "Resolved")
add_issue("Verdant Bio", None,
          "Stored Ownership% column (25% -> 19.2% -> 7.3% -> 5.2%, "
          "monotonically decreasing) diverges from amount_raised/post_money "
          "recompute (25% -> 23.1% -> 18.2% -> 14.9%).",
          "Recomputed ownership_pct_new_investor per the standard rule for "
          "all rows. Stored values represent Engine's own cumulative "
          "diluting fund position, not each round's new-money %, and were "
          "preserved separately in ownership_pct_fund_position.",
          "Resolved")
add_issue("Verdant Bio", "Series C",
          "Series C carries a firm date (4/1/2025) and full deal terms in "
          "the tracker, but is not confirmed closed.",
          "Set round_status='Planned', date_closed=NULL. Expected/target "
          "close date (2025-04-01, reflecting the 'pushed to Q2 2025' "
          "update in Portfolio Notes) recorded in source_note instead of "
          "date_closed.",
          "Resolved", round_id=verdant_series_c_id)
add_issue("Verdant Bio", "Series C",
          "Portfolio Notes: 'Verdant Bio Series C round pushed back a "
          "quarter, expected close now Q2 2025 not Q1'.",
          "Cross-checked against tracker date (4/1/2025, which is Q2) -- "
          "already reflects this update, no date change needed. (Status "
          "handled separately, see other Series C entry.)",
          "Resolved", round_id=verdant_series_c_id)

# ---------------------------------------------------------------------------
# Fathom Analytics
# ---------------------------------------------------------------------------
# This is the sheet whose Pre-Money/Post-Money columns are swapped relative
# to the others -- reading by header name (not position) in extract_rows()
# is what makes that a non-issue here.
raw = extract_rows("Fathom Analytics", header_row=1, data_rows=range(2, 6))
orders = assign_round_orders(raw, lambda rec: (rec["Round"], rec["Date"]))
fathom_pre_seed_id = None
fathom_series_b_id = None
for rec, order in zip(raw, orders):
    name = rec["Round"]
    date_closed = parse_date(rec["Date"])
    amount, pre, post = rec["Amount Raised"], rec.get("Pre-Money"), rec.get("Post-Money")
    price = rec.get("Price/Share")
    rtype = classify_round_type(name, pre)

    if name == "Pre-Seed":
        # A separately-supplied note mentions a possibly-duplicate/conflicting
        # $1.2M SAFE dated the same day as this $750K row; the tracker's own
        # figure is kept as-is (not overwritten) but flagged for review.
        fathom_pre_seed_id = add_round(
                   "Fathom Analytics", name, order, date_closed, "Closed", rtype,
                   amount, pre, post, price, None, pct(amount, post), None,
                   "needs_review", 0,
                   "Tracker's existing $750K figure retained; Portfolio "
                   "Notes mentions an unconfirmed 'new SAFE $1.2M dated "
                   "1/10/22' referencing the same date -- possible "
                   "duplicate/conflict, not applied. See data_quality_log.")
    else:
        rid = add_round("Fathom Analytics", name, order, date_closed, "Closed", rtype,
                   amount, pre, post, price, None, pct(amount, post), None,
                   "confirmed", 0, "Tracker sheet, reconciled, no adjustments.")
        if name == "Series B":
            fathom_series_b_id = rid

add_issue("Fathom Analytics", "Pre-Seed",
          "Portfolio Notes mentions an unconfirmed 'new SAFE, $1.2M, dated "
          "1/10/22' -- same date as the tracker's existing Pre-Seed row, "
          "which shows $750K raised. Possible duplicate or conflicting "
          "entry for the same event.",
          "Kept the tracker's existing $750K figure (not overwritten by "
          "the unconfirmed note); flagged source_confidence='needs_review' "
          "pending deal-team confirmation.",
          "Open", round_id=fathom_pre_seed_id)
add_issue("Fathom Analytics", "Series B",
          "Portfolio Notes: 'Fathom Analytics Inc. Series B - closed June "
          "2025, $22M raised, post money ~$92M'.",
          "Cross-checked against tracker (6/18/2025, $22M raised, $92M "
          "post) -- matches, no change needed.",
          "Resolved", round_id=fathom_series_b_id)

# ---------------------------------------------------------------------------
# Ridgeline Materials ($000s -> multiply by 1000)
# ---------------------------------------------------------------------------
# This sheet's own footnote states every dollar figure is quoted in
# thousands, unlike every other sheet -- all dollar fields below get
# multiplied by RIDGELINE_MULT before being inserted.
RIDGELINE_MULT = 1000
raw = extract_rows("Ridgeline Materials", header_row=1, data_rows=range(2, 7))
orders = assign_round_orders(raw, lambda rec: (rec["Round"], rec["Date"]))
ridgeline_series_b_id = None
for rec, order in zip(raw, orders):
    name = rec["Round"]
    date_closed = parse_date(rec["Date"])
    amount = rec["Amount Raised"]
    pre, post = rec.get("Pre-Money"), rec.get("Post-Money")

    note = "Unit conversion: source figures in $000s, multiplied by 1,000."
    if name == "Series B" and amount < 0:
        # Sheet footnote and Portfolio Notes both confirm this negative
        # amount is a data-entry typo (should be positive), not a real
        # negative raise -- correct the sign before converting units.
        amount = abs(amount)
        note += (" Amount Raised corrected from -22,000 to +22,000 (typo "
                 "confirmed by sheet footnote and Portfolio Notes).")

    # Apply the $000s -> $ conversion to every dollar-denominated field.
    amount = amount * RIDGELINE_MULT if amount is not None else None
    pre = pre * RIDGELINE_MULT if pre is not None else None
    post = post * RIDGELINE_MULT if post is not None else None
    rtype = classify_round_type(name, pre)

    rid = add_round("Ridgeline Materials", name, order, date_closed, "Closed", rtype,
               amount, pre, post, None, None, pct(amount, post), None,
               "confirmed", 0, note)
    if name == "Series B":
        ridgeline_series_b_id = rid

add_issue("Ridgeline Materials", None,
          "Sheet footnote states all dollar figures are in $000s, differing "
          "from every other sheet's raw-dollar convention.",
          "Multiplied all Ridgeline dollar fields (amount_raised_usd, "
          "pre_money_usd, post_money_usd) by 1,000 before inserting.",
          "Resolved")
add_issue("Ridgeline Materials", "Series B",
          "Amount Raised entered as -22,000 (in $000s) -- both the sheet's "
          "own footnote and Portfolio Notes confirm this is a typo.",
          "Corrected to +22,000 ($22M after unit conversion) before insert.",
          "Resolved", round_id=ridgeline_series_b_id)

# ---------------------------------------------------------------------------
# Halcyon Health
# ---------------------------------------------------------------------------
raw = extract_rows("Halcyon Health", header_row=1, data_rows=range(2, 6))
# Drop the broken $9.0M Series A row (shares=0, #DIV/0!) before assigning order,
# so it never becomes an inserted row or consumes an order slot.
raw = [r for r in raw if not (r["Round"] == "Series A" and r.get("Shares Post-Round") == 0)]
orders = assign_round_orders(raw, lambda rec: (rec["Round"], rec["Date"]))
halcyon_series_a_id = None
halcyon_series_b_id = None
for rec, order in zip(raw, orders):
    name = rec["Round"]
    date_closed = parse_date(rec["Date"])
    amount, pre, post = rec["Amount Raised"], rec.get("Pre-Money"), rec.get("Post-Money")
    shares = rec.get("Shares Post-Round")
    # Treat a share count of exactly 0 the same as missing -- it's a broken
    # formula artifact, not a real "zero shares" fact.
    shares = None if shares in (0, None) else shares
    rtype = classify_round_type(name, pre)

    if name == "Series A":
        # Only the $9.5M/15.2M-share row is kept (the other conflicting
        # Series A row was already filtered out above for having 0 shares);
        # still flagged as needs_review since the two rows were never
        # reconciled against each other in the source.
        halcyon_series_a_id = add_round(
                   "Halcyon Health", name, order, date_closed, "Closed", rtype,
                   amount, pre, post, None, shares, pct(amount, post), None,
                   "needs_review", 1,
                   "Two Series A rows existed in source (sheet footnote: "
                   "'not yet reconciled'); this $9.5M/15.2M-share row "
                   "retained as the only one with usable share data. "
                   "Alternate unconfirmed figure: $9.0M raised, "
                   "shares_post_round and Ownership% broken (#DIV/0!) in "
                   "source. Pending finance confirmation per Portfolio Notes.")
    elif name == "Series B":
        # Share count is broken in the source for this round too, but the
        # dollar figures are independently reliable -- left shares NULL
        # rather than guessing a value.
        halcyon_series_b_id = add_round(
                   "Halcyon Health", name, order, date_closed, "Closed", rtype,
                   amount, pre, post, None, None, pct(amount, post), None,
                   "confirmed", 0,
                   "shares_post_round unavailable in source (#DIV/0!); left "
                   "NULL at extraction rather than estimated. Amount/pre/"
                   "post-money figures reliable independent of the broken "
                   "share count. price_per_share/shares_post_round later "
                   "back-computed by chaining from Series A's confirmed "
                   "Shares Post -- see data_quality_log.")
    else:
        add_round("Halcyon Health", name, order, date_closed, "Closed", rtype,
                   amount, pre, post, None, shares, pct(amount, post), None,
                   "confirmed", 0, "Tracker sheet, reconciled, no adjustments.")

add_issue("Halcyon Health", "Series A",
          "Two conflicting Series A rows in source: $9.0M raised "
          "(shares_post_round=0, #DIV/0! ownership -- broken formula "
          "artifact) vs. $9.5M raised (complete, 15.2M shares). Sheet "
          "footnote and Portfolio Notes both flag as unreconciled, pending "
          "finance confirmation.",
          "Inserted only the $9.5M row (the only one with usable share "
          "data), flagged source_confidence='needs_review', is_estimate=1. "
          "$9.0M figure preserved here for reference, not inserted as a row.",
          "Open", round_id=halcyon_series_a_id)
add_issue("Halcyon Health", "Series B",
          "shares_post_round was 0/#DIV/0! in source (broken formula, no "
          "real share count available).",
          "Left shares_post_round and price_per_share NULL rather than "
          "estimate a value. ownership_pct_new_investor still computed "
          "from amount_raised_usd/post_money_usd, both of which are reliable.",
          "Resolved", round_id=halcyon_series_b_id)

# ---------------------------------------------------------------------------
# Dataset-wide issues (span multiple companies -> single consolidated row)
# ---------------------------------------------------------------------------
# Logged against the synthetic "All Companies" company_name (see
# load_quality_log in app.py, which surfaces these on every company's page).
add_issue("All Companies", None,
          "Every sheet's stored Ownership% column uses a different, "
          "unreliable basis (e.g. Halcyon's equals prior_shares/"
          "total_shares_post_round, a cap-table stat, not deal ownership; "
          "two Halcyon rows are outright #DIV/0!).",
          "Recomputed ownership_pct_new_investor for every row as "
          "amount_raised_usd / post_money_usd; no sheet's stored "
          "Ownership% column was used directly (Verdant's stored values "
          "were preserved separately as ownership_pct_fund_position, see "
          "separate entry).",
          "Resolved")
add_issue("All Companies", None,
          "Footnote/context text rows are embedded directly inside the "
          "data ranges on several sheets (Nimbus row 11, Ridgeline row 8, "
          "Halcyon row 7, Verdant row 17), which if not filtered would "
          "become garbage financing_rounds rows with null amounts.",
          "Rows without a valid Amount Raised value were excluded from "
          "extraction; footnote text was reviewed and folded into the "
          "relevant data_quality_log entries above as context rather than "
          "inserted as rows.",
          "Resolved")

# ---------------------------------------------------------------------------
# Backfill computable fields (Price/Share, Shares Post, Own% New Inv.)
# ---------------------------------------------------------------------------
# One pass over every company's already-inserted rows, walking each
# company's rounds chronologically (round_order) since later rounds' Rule 4
# depends on an earlier round's confirmed Shares Post. See README "Data
# Computation & NULL Handling" for the plain-language version of rules 1-9
# implemented below. Rule 1 (SAFE/Note rows left NULL) needs no code here --
# extraction never populates pre/post/price/shares for those rows in the
# first place.

def backfill_computed_fields():
    by_company = defaultdict(list)
    for idx, r in enumerate(rounds):
        by_company[r["company_name"]].append(idx)

    for company, idxs in by_company.items():
        # --- Rule 8: flag same-round_order rows with differing terms; never
        # compute onto them, and don't re-flag a conflict this file's own
        # per-company block already flagged (e.g. Nimbus Series A2). -------
        order_groups = defaultdict(list)
        for idx in idxs:
            order_groups[rounds[idx]["round_order"]].append(idx)
        conflict_idxs = set()
        compare_keys = ("amount_raised_usd", "pre_money_usd", "post_money_usd",
                         "price_per_share", "shares_post_round")
        for order, group in order_groups.items():
            if len(group) < 2:
                continue
            first = rounds[group[0]]
            if any(rounds[i][k] != first[k] for i in group[1:] for k in compare_keys):
                conflict_idxs.update(group)
                if not all(rounds[i]["source_confidence"] == "needs_review" for i in group):
                    add_issue(company, first["round_name"],
                        f"Rule 8 duplicate scan: {len(group)} rows share "
                        f"round_order={order} ('{first['round_name']}') with "
                        f"differing terms.",
                        "Left all rows as-is; no auto-computed fields applied "
                        "to this round_order -- flagged for manual resolution.",
                        "Open", round_id=group[0] + 1)

        # --- Rules 2-7, walked chronologically ----------------------------
        confirmed_shares = None  # most recent known shares_post_round so far
        sanity_fail_idxs = set()
        for idx in idxs:
            r = rounds[idx]
            round_id = idx + 1
            filled = []

            if idx in conflict_idxs:
                if r["shares_post_round"] is not None:
                    confirmed_shares = r["shares_post_round"]
                continue

            if r["round_type"] == "Priced Equity":
                pre, post = r["pre_money_usd"], r["post_money_usd"]
                price, shares = r["price_per_share"], r["shares_post_round"]
                amount = r["amount_raised_usd"]

                if price is not None and shares is None and post is not None:
                    # Rule 2: Shares Post = Post-Money / Price-Share
                    r["shares_post_round"] = round(post / price)
                    filled.append("shares_post_round (Rule 2)")

                elif shares is not None and price is None and post is not None:
                    # Rule 3: Price-Share = Post-Money / Shares Post
                    r["price_per_share"] = round(post / shares, 2)
                    filled.append("price_per_share (Rule 3)")

                elif price is None and shares is None:
                    # Rule 4: chain off the prior round's confirmed Shares Post.
                    if (confirmed_shares is not None and pre is not None
                            and post is not None and amount is not None):
                        price_unrounded = pre / confirmed_shares
                        shares_unrounded = post / price_unrounded
                        new_shares_unrounded = amount / price_unrounded
                        reconciled = confirmed_shares + new_shares_unrounded
                        pct_diff = abs(reconciled - shares_unrounded) / shares_unrounded
                        if pct_diff <= 0.01:
                            r["price_per_share"] = round(price_unrounded, 2)
                            r["shares_post_round"] = round(shares_unrounded)
                            filled.append(
                                f"price_per_share & shares_post_round (Rule 4, "
                                f"chained from prior Shares Post = "
                                f"{confirmed_shares:,.0f})")
                        else:
                            sanity_fail_idxs.add(idx)
                            add_issue(company, r["round_name"],
                                f"Rule 4 sanity check failed: chaining Price/Share "
                                f"from Pre-Money / prior Shares Post "
                                f"({confirmed_shares:,.0f}) implies "
                                f"{shares_unrounded:,.0f} Shares Post, but prior "
                                f"Shares Post + (Amount Raised / Price/Share) = "
                                f"{reconciled:,.0f} -- a {pct_diff:.2%} mismatch, "
                                f"over the 1% tolerance.",
                                "Left price_per_share and shares_post_round NULL "
                                "rather than guessing; flagged for manual review.",
                                "Open", round_id=round_id)
                    # else Rule 4e: no prior confirmed Shares Post to chain
                    # from -- leave both NULL, no per-row flag (expected gap,
                    # covered by the company-level summary note below).

            if r["shares_post_round"] is not None:
                confirmed_shares = r["shares_post_round"]

            # --- Rule 5: Own% (New Inv.) = Amount Raised / Post-Money ------
            amount, post = r["amount_raised_usd"], r["post_money_usd"]
            if amount is not None and post is not None:
                computed_own = amount / post
                if r["ownership_pct_new_investor"] is None:
                    r["ownership_pct_new_investor"] = computed_own
                    filled.append("ownership_pct_new_investor (Rule 5)")
                elif abs(computed_own - r["ownership_pct_new_investor"]) * 100 > 0.5:
                    add_issue(company, r["round_name"],
                        f"Rule 5 conflict: Amount Raised / Post-Money "
                        f"({computed_own:.2%}) differs from the stored "
                        f"ownership_pct_new_investor "
                        f"({r['ownership_pct_new_investor']:.2%}) by more than "
                        f"0.5 percentage points.",
                        "Did not overwrite the existing stored value; flagged "
                        "for manual reconciliation instead.",
                        "Open", round_id=round_id)

            # --- Rule 7: unclosed rounds -> anything filled here is a
            # projection, not a confirmed fact. -----------------------------
            if filled and r["round_status"] == "Planned":
                r["source_confidence"] = "needs_review"
                r["is_estimate"] = 1

            if filled:
                label = ("projected/unconfirmed, round is Planned and not yet "
                          "closed" if r["round_status"] == "Planned"
                          else "computed from this closed round's own figures")
                add_issue(company, r["round_name"],
                    f"Backfill: {', '.join(filled)} had no source value.",
                    f"Auto-computed per the documented rules ({label}); see "
                    f"README 'Data Computation & NULL Handling' for the "
                    f"formulas used.",
                    "Resolved", round_id=round_id)

        # --- Company-level summary for the Rule 4e "expected gap" case: a
        # priced round with no fillable prior Shares Post anywhere in the
        # company's history (so nothing above ever flagged or filled it). --
        still_missing = [
            rounds[i]["round_name"] for i in idxs
            if i not in conflict_idxs and i not in sanity_fail_idxs
            and rounds[i]["round_type"] == "Priced Equity"
            and rounds[i]["price_per_share"] is None
            and rounds[i]["shares_post_round"] is None
        ]
        if still_missing:
            add_issue(company, None,
                f"Backfill (Rule 4e): {', '.join(still_missing)} priced "
                f"round(s) have no Price/Share or Shares Post recorded, and "
                f"no earlier round for {company} has a confirmed Shares Post "
                f"to chain a per-share price from.",
                "Left NULL rather than estimated; Price/Share and Shares "
                "Post need either a direct source value or a prior round's "
                "confirmed Shares Post to derive from (see README).",
                "Resolved")


backfill_computed_fields()

# ---------------------------------------------------------------------------
# Write MySQL-flavored SQL dump (for DataGrip, and executed against MySQL below)
# ---------------------------------------------------------------------------
# portfolio.sql is a hand-generated MySQL-compatible schema + data dump built
# straight from the `rounds`/`log` accumulators above. It's both the
# version-controlled, human-diffable artifact you'd open in DataGrip, and
# the exact script this script itself runs against MySQL next.

def sql_str(value):
    """Render a Python value as a MySQL literal: quoted/escaped string, a
    bare number, or NULL. Both '' and \\\\ are escaped since MySQL's default
    sql_mode treats backslash as an escape character."""
    if value is None:
        return "NULL"
    if isinstance(value, (int, float)):
        return repr(value)
    escaped = str(value).replace("\\", "\\\\").replace("'", "''")
    return f"'{escaped}'"


MYSQL_DATABASE = "portfolio"

# DROP DATABASE + CREATE DATABASE up front means every individual DROP TABLE
# below is guaranteed to be a no-op (nothing exists yet in the fresh
# database) -- this is what actually makes the whole script safe to re-run,
# rather than relying on FK-aware DROP TABLE ordering.
MYSQL_SCHEMA = f"""
DROP DATABASE IF EXISTS {MYSQL_DATABASE};
CREATE DATABASE {MYSQL_DATABASE};
USE {MYSQL_DATABASE};

DROP TABLE IF EXISTS companies;
CREATE TABLE IF NOT EXISTS companies (
    company_id INT AUTO_INCREMENT PRIMARY KEY,
    company_name VARCHAR(255) UNIQUE NOT NULL
);

DROP TABLE IF EXISTS financing_rounds;
CREATE TABLE IF NOT EXISTS financing_rounds (
    round_id INT AUTO_INCREMENT PRIMARY KEY,
    company_id INT NOT NULL,
    round_name VARCHAR(255) NOT NULL,
    round_order INT NOT NULL,
    date_closed DATE,
    round_status VARCHAR(20) NOT NULL CHECK (round_status IN ('Closed', 'Planned')),
    round_type VARCHAR(30) NOT NULL CHECK (round_type IN ('SAFE', 'Convertible Note', 'Priced Equity')),
    amount_raised_usd DOUBLE,
    pre_money_usd DOUBLE,
    post_money_usd DOUBLE,
    price_per_share DOUBLE,
    shares_post_round DOUBLE,
    ownership_pct_new_investor DOUBLE,
    ownership_pct_fund_position DOUBLE,
    source_confidence VARCHAR(20) NOT NULL CHECK (source_confidence IN ('confirmed', 'needs_review')),
    is_estimate TINYINT NOT NULL CHECK (is_estimate IN (0, 1)),
    source_note TEXT,
    created_at VARCHAR(32) NOT NULL,
    updated_at VARCHAR(32) NOT NULL,
    CONSTRAINT fk_financingRoundsCompany FOREIGN KEY (company_id) REFERENCES companies (company_id)
        ON UPDATE CASCADE
        ON DELETE CASCADE
);

DROP TABLE IF EXISTS data_quality_log;
CREATE TABLE IF NOT EXISTS data_quality_log (
    issue_id INT AUTO_INCREMENT PRIMARY KEY,
    company_name VARCHAR(255) NOT NULL,
    round_name VARCHAR(255),
    company_id INT NULL,
    round_id INT NULL,
    issue TEXT NOT NULL,
    resolution TEXT NOT NULL,
    status VARCHAR(20) NOT NULL CHECK (status IN ('Resolved', 'Open')),
    logged_at VARCHAR(32) NOT NULL,
    CONSTRAINT fk_dataQualityLogCompany FOREIGN KEY (company_id) REFERENCES companies (company_id)
        ON UPDATE CASCADE
        ON DELETE SET NULL,
    CONSTRAINT fk_dataQualityLogRound FOREIGN KEY (round_id) REFERENCES financing_rounds (round_id)
        ON UPDATE CASCADE
        ON DELETE SET NULL
);
""".strip()

with open(SQL_DUMP_PATH, "w", encoding="utf-8") as f:
    f.write("-- MySQL dump for the \"portfolio\" database -- open this in\n")
    f.write("-- DataGrip, or it's run directly against MySQL by this script.\n")
    f.write("-- Safe to re-run: DROP DATABASE + CREATE DATABASE up top means\n")
    f.write("-- the whole script always starts from a clean slate, so the\n")
    f.write("-- per-table DROP/CREATE and FK constraints below never conflict\n")
    f.write("-- with a prior run.\n\n")
    f.write(MYSQL_SCHEMA + "\n\n")

    f.write("INSERT INTO companies (company_id, company_name) VALUES\n")
    f.write(",\n".join(
        f"    ({idx}, {sql_str(name)})" for idx, name in enumerate(COMPANIES, start=1)
    ) + ";\n\n")

    round_columns = (
        "round_id, company_id, round_name, round_order, date_closed, round_status, "
        "round_type, amount_raised_usd, pre_money_usd, post_money_usd, price_per_share, "
        "shares_post_round, ownership_pct_new_investor, ownership_pct_fund_position, "
        "source_confidence, is_estimate, source_note, created_at, updated_at"
    )
    f.write(f"INSERT INTO financing_rounds ({round_columns}) VALUES\n")
    f.write(",\n".join(
        "    (" + ", ".join([
            str(round_id),
            str(company_id[r["company_name"]]),
            sql_str(r["round_name"]), str(r["round_order"]), sql_str(r["date_closed"]),
            sql_str(r["round_status"]), sql_str(r["round_type"]),
            sql_str(r["amount_raised_usd"]), sql_str(r["pre_money_usd"]), sql_str(r["post_money_usd"]),
            sql_str(r["price_per_share"]), sql_str(r["shares_post_round"]),
            sql_str(r["ownership_pct_new_investor"]), sql_str(r["ownership_pct_fund_position"]),
            sql_str(r["source_confidence"]), str(r["is_estimate"]), sql_str(r["source_note"]),
            sql_str(NOW), sql_str(NOW),
        ]) + ")"
        for round_id, r in enumerate(rounds, start=1)
    ) + ";\n\n")

    issue_columns = (
        "issue_id, company_name, round_name, company_id, round_id, "
        "issue, resolution, status, logged_at"
    )
    f.write(f"INSERT INTO data_quality_log ({issue_columns}) VALUES\n")
    f.write(",\n".join(
        "    (" + ", ".join([
            str(issue_id), sql_str(i["company_name"]), sql_str(i["round_name"]),
            sql_str(i["company_id"]), sql_str(i["round_id"]),
            sql_str(i["issue"]), sql_str(i["resolution"]), sql_str(i["status"]), sql_str(NOW),
        ]) + ")"
        for issue_id, i in enumerate(log, start=1)
    ) + ";\n")

# ---------------------------------------------------------------------------
# Execute portfolio.sql against MySQL
# ---------------------------------------------------------------------------
# Run the exact file just written -- not a re-derivation from the Python
# data -- so the sanity checks and CSV export below are validating the real,
# shipped artifact (catching e.g. a SQL syntax slip) rather than the
# pre-SQL Python structures. CLIENT.MULTI_STATEMENTS lets one cursor.execute()
# run the whole multi-statement script; connecting with no database= (the
# script itself does DROP DATABASE/CREATE DATABASE/USE) avoids "can't drop
# the database you're currently connected to" ordering issues.
conn = pymysql.connect(
    host=MYSQL_CREDS["host"], port=int(MYSQL_CREDS.get("port", 3306)),
    user=MYSQL_CREDS["user"], password=MYSQL_CREDS["password"],
    client_flag=CLIENT.MULTI_STATEMENTS,
)
cur = conn.cursor()
with open(SQL_DUMP_PATH, "r", encoding="utf-8") as f:
    cur.execute(f.read())
while cur.nextset():
    pass
conn.commit()

# ---------------------------------------------------------------------------
# Sanity checks
# ---------------------------------------------------------------------------
# Basic post-load assertions, printed rather than raised, so a failed check
# is visible in the run output without stopping the script from finishing
# the CSV export below. Queried against the MySQL database that was just
# built (the same one app.py reads), not a separate SQLite copy.
print("=" * 80)
print("SANITY CHECKS")
print("=" * 80)

# Every Closed round should have an amount raised -- a NULL here would mean
# a real financing event lost its amount somewhere in extraction.
cur.execute("""
    SELECT c.company_name, fr.round_name FROM financing_rounds fr
    JOIN companies c ON c.company_id = fr.company_id
    WHERE fr.round_status = 'Closed' AND fr.amount_raised_usd IS NULL
""")
missing_amount = cur.fetchall()
print(f"Closed rounds missing amount_raised_usd: {len(missing_amount)}")
for row in missing_amount:
    print("  FAIL:", row)

# ownership_pct_new_investor is a fraction of the company (0-100%); anything
# outside [0, 1] would indicate a unit-conversion or sign error upstream.
cur.execute("""
    SELECT c.company_name, fr.round_name, fr.ownership_pct_new_investor
    FROM financing_rounds fr
    JOIN companies c ON c.company_id = fr.company_id
    WHERE fr.ownership_pct_new_investor < 0 OR fr.ownership_pct_new_investor > 1.0
""")
bad_ownership = cur.fetchall()
print(f"ownership_pct_new_investor out of [0,1] range: {len(bad_ownership)}")
for row in bad_ownership:
    print("  FAIL:", row)

if not missing_amount and not bad_ownership:
    print("All sanity checks passed.")

# ---------------------------------------------------------------------------
# Print data_quality_log for review
# ---------------------------------------------------------------------------
cur.execute("""
    SELECT issue_id, company_name, round_name, issue, resolution, status, logged_at
    FROM data_quality_log ORDER BY issue_id
""")
rows_out = cur.fetchall()

# Human-readable dump of every logged issue to stdout, for a quick end-of-run review.
print("\n" + "=" * 80)
print(f"DATA QUALITY LOG ({len(rows_out)} entries)")
print("=" * 80)
for row in rows_out:
    issue_id, comp, rname, issue, resolution, status, logged_at = row
    print(f"\n[{issue_id}] {comp}" + (f" / {rname}" if rname else "") + f" -- {status}")
    print(f"  ISSUE:      {issue}")
    print(f"  RESOLUTION: {resolution}")

open_count = sum(1 for r in rows_out if r[5] == "Open")
resolved_count = sum(1 for r in rows_out if r[5] == "Resolved")
print(f"\nTotal: {len(rows_out)}  |  Resolved: {resolved_count}  |  Open: {open_count}")
print(f"\nMySQL database '{MYSQL_CREDS['database']}' rebuilt from {SQL_DUMP_PATH}")

conn.close()
